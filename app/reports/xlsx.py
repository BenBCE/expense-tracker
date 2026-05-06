"""Excel report generator."""

from __future__ import annotations

from collections import defaultdict
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.reports.types import ReportData

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CURRENCY_FMT = '#,##0.00;[Red]-#,##0.00'

EXPENSE_COLUMNS = [
    ("Date", 12),
    ("Vendor", 28),
    ("Category", 14),
    ("Subtotal", 12),
    ("VAT", 12),
    ("Total", 12),
    ("Currency", 10),
    ("Note", 32),
    ("Receipt File", 36),
]

SUMMARY_COLUMNS = [
    ("Category", 16),
    ("Currency", 10),
    ("Total", 14),
    ("Count", 8),
]


def _write_header(ws, columns) -> None:  # type: ignore[no-untyped-def]
    for idx, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def build_xlsx(data: ReportData, out_path: Path) -> Path:
    wb = Workbook()
    expenses = wb.active
    expenses.title = "Expenses"
    _write_header(expenses, EXPENSE_COLUMNS)

    for row_idx, r in enumerate(data.rows, start=2):
        expenses.cell(row=row_idx, column=1, value=r.date or "")
        expenses.cell(row=row_idx, column=2, value=r.vendor or "")
        expenses.cell(row=row_idx, column=3, value=r.category or "")
        expenses.cell(
            row=row_idx,
            column=4,
            value=float(r.subtotal) if r.subtotal is not None else None,
        ).number_format = CURRENCY_FMT
        expenses.cell(
            row=row_idx,
            column=5,
            value=float(r.vat) if r.vat is not None else None,
        ).number_format = CURRENCY_FMT
        expenses.cell(
            row=row_idx,
            column=6,
            value=float(r.total) if r.total is not None else None,
        ).number_format = CURRENCY_FMT
        expenses.cell(row=row_idx, column=7, value=r.currency or "")
        expenses.cell(row=row_idx, column=8, value=r.note or "")
        expenses.cell(row=row_idx, column=9, value=r.s3_key)

    summary = wb.create_sheet("Summary")
    _write_header(summary, SUMMARY_COLUMNS)

    by_cat: dict[tuple[str, str], list[Decimal]] = defaultdict(list)
    for r in data.rows:
        if r.total is None:
            continue
        by_cat[(r.category or "other", r.currency or "")].append(Decimal(str(r.total)))

    row_idx = 2
    grand_by_currency: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for (cat, cur), totals in sorted(by_cat.items()):
        cat_total = sum(totals, Decimal("0"))
        summary.cell(row=row_idx, column=1, value=cat)
        summary.cell(row=row_idx, column=2, value=cur)
        c = summary.cell(row=row_idx, column=3, value=float(cat_total))
        c.number_format = CURRENCY_FMT
        summary.cell(row=row_idx, column=4, value=len(totals))
        grand_by_currency[cur] += cat_total
        row_idx += 1

    row_idx += 1
    for cur, total in sorted(grand_by_currency.items()):
        summary.cell(row=row_idx, column=1, value="GRAND TOTAL").font = Font(bold=True)
        summary.cell(row=row_idx, column=2, value=cur).font = Font(bold=True)
        c = summary.cell(row=row_idx, column=3, value=float(total))
        c.number_format = CURRENCY_FMT
        c.font = Font(bold=True)
        row_idx += 1

    wb.save(out_path)
    return out_path
